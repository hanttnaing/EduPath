import csv
import shutil
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl


CSV_PATH = Path(
    "data/cleaned/"
    "japan_programs_intake_deadline_enriched.csv"
)

WORKBOOK_PATH = Path(
    "data/sample/06_full_dataset.xlsx"
)

SHEET_NAME = "programs"


EXPECTED_HEADERS = [
    "program_id",
    "university_id",
    "program_name",
    "field_of_study",
    "degree_level",
    "duration_years",
    "study_mode",
    "language_of_instruction",
    "tuition_fee",
    "tuition_currency",
    "tuition_period",
    "minimum_gpa",
    "gpa_scale",
    "ielts_requirement",
    "toefl_requirement",
    "intake",
    "application_deadline",
    "program_url",
    "collected_at",
    "last_verified_at",
    "freshness_status",
]


EXPECTED_JAPAN_PROGRAMS = 36


def convert_value(
    header: str,
    value: str,
):
    value = value.strip()

    if value == "":
        return None

    # ---------------------------------
    # Integer-like fields
    # ---------------------------------

    if header in {
        "tuition_fee",
    }:
        try:
            return int(
                float(value)
            )
        except ValueError:
            return value

    # ---------------------------------
    # Decimal / numeric fields
    # ---------------------------------

    if header in {
        "duration_years",
        "minimum_gpa",
        "gpa_scale",
        "ielts_requirement",
        "toefl_requirement",
    }:
        try:
            number = float(value)

            if number.is_integer():
                return int(number)

            return number

        except ValueError:
            return value

    # ---------------------------------
    # Date fields
    # ---------------------------------

    if header in {
        "application_deadline",
        "collected_at",
        "last_verified_at",
    }:
        try:
            return date.fromisoformat(
                value
            )
        except ValueError:
            return value

    return value


def is_japan_program(
    program_id,
    university_id,
) -> bool:
    program_id = (
        str(program_id).strip()
        if program_id is not None
        else ""
    )

    university_id = (
        str(university_id).strip()
        if university_id is not None
        else ""
    )

    return (
        program_id.startswith(
            "prog_jp_"
        )
        or university_id.startswith(
            "uni_jp_"
        )
    )


def main() -> None:
    # ---------------------------------
    # 1. Check required files
    # ---------------------------------

    if not CSV_PATH.exists():
        raise FileNotFoundError(
            f"CSV file not found: "
            f"{CSV_PATH}"
        )

    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: "
            f"{WORKBOOK_PATH}"
        )

    # ---------------------------------
    # 2. Read Japan program dataset
    # ---------------------------------

    with CSV_PATH.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as csv_file:

        reader = csv.DictReader(
            csv_file
        )

        csv_headers = reader.fieldnames

        if csv_headers != EXPECTED_HEADERS:
            raise ValueError(
                "Japan program CSV headers "
                "do not match the expected "
                "21-column program schema."
            )

        japan_rows = list(reader)

    print(
        f"Japan program CSV records: "
        f"{len(japan_rows)}"
    )

    if (
        len(japan_rows)
        != EXPECTED_JAPAN_PROGRAMS
    ):
        raise ValueError(
            "Expected exactly "
            f"{EXPECTED_JAPAN_PROGRAMS} "
            "Japan program records."
        )

    # ---------------------------------
    # 3. Validate IDs before touching
    #    the workbook
    # ---------------------------------

    source_program_ids = [
        row["program_id"].strip()
        for row in japan_rows
    ]

    if (
        len(source_program_ids)
        != len(set(source_program_ids))
    ):
        raise ValueError(
            "Duplicate program_id found "
            "in Japan source CSV."
        )

    invalid_country_rows = [
        row["program_id"]
        for row in japan_rows
        if not row[
            "program_id"
        ].startswith("prog_jp_")
        or not row[
            "university_id"
        ].startswith("uni_jp_")
    ]

    if invalid_country_rows:
        raise ValueError(
            "Non-Japan ID detected in "
            "Japan program dataset: "
            + ", ".join(
                invalid_country_rows
            )
        )

    # ---------------------------------
    # 4. Safety backup
    # ---------------------------------

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    backup_path = (
        WORKBOOK_PATH.with_name(
            "06_full_dataset_before_"
            f"japan_program_load_"
            f"{timestamp}.xlsx"
        )
    )

    shutil.copy2(
        WORKBOOK_PATH,
        backup_path,
    )

    print(
        f"Backup created: "
        f"{backup_path}"
    )

    # ---------------------------------
    # 5. Open workbook
    # ---------------------------------

    workbook = openpyxl.load_workbook(
        WORKBOOK_PATH
    )

    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()

        raise ValueError(
            f"Sheet '{SHEET_NAME}' "
            "was not found."
        )

    worksheet = workbook[
        SHEET_NAME
    ]

    # ---------------------------------
    # 6. Validate sheet headers
    # ---------------------------------

    workbook_headers = [
        worksheet.cell(
            row=1,
            column=column_number,
        ).value
        for column_number in range(
            1,
            len(EXPECTED_HEADERS) + 1,
        )
    ]

    if workbook_headers != EXPECTED_HEADERS:
        workbook.close()

        raise ValueError(
            "Programs sheet headers "
            "do not match the expected "
            "21-column schema."
        )

    print(
        "Workbook program schema: OK"
    )

    # ---------------------------------
    # 7. Preserve row style
    # ---------------------------------

    template_styles = {}

    if worksheet.max_row >= 2:
        for column_number in range(
            1,
            len(EXPECTED_HEADERS) + 1,
        ):
            cell = worksheet.cell(
                row=2,
                column=column_number,
            )

            template_styles[
                column_number
            ] = {
                "font": copy(
                    cell.font
                ),
                "fill": copy(
                    cell.fill
                ),
                "border": copy(
                    cell.border
                ),
                "alignment": copy(
                    cell.alignment
                ),
                "protection": copy(
                    cell.protection
                ),
                "number_format": (
                    cell.number_format
                ),
            }

    template_row_height = None

    if worksheet.max_row >= 2:
        template_row_height = (
            worksheet.row_dimensions[
                2
            ].height
        )

    # ---------------------------------
    # 8. Preserve NON-Japan rows
    # ---------------------------------

    existing_non_japan_rows = []

    if worksheet.max_row >= 2:
        for row_number in range(
            2,
            worksheet.max_row + 1,
        ):
            values = [
                worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value
                for column_number in range(
                    1,
                    len(EXPECTED_HEADERS) + 1,
                )
            ]

            # Skip completely blank rows
            if not any(
                value is not None
                for value in values
            ):
                continue

            program_id = values[0]
            university_id = values[1]

            if is_japan_program(
                program_id,
                university_id,
            ):
                continue

            existing_non_japan_rows.append(
                values
            )

    print(
        "Existing non-Japan programs "
        f"preserved: "
        f"{len(existing_non_japan_rows)}"
    )

    # ---------------------------------
    # 9. Clear current data rows
    # ---------------------------------

    if worksheet.max_row > 1:
        worksheet.delete_rows(
            2,
            worksheet.max_row - 1,
        )

    # ---------------------------------
    # 10. Write preserved non-Japan
    #     rows first
    # ---------------------------------

    next_row = 2

    for values in existing_non_japan_rows:
        for column_number, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=next_row,
                column=column_number,
            )

            cell.value = value

            if (
                column_number
                in template_styles
            ):
                style = template_styles[
                    column_number
                ]

                cell.font = copy(
                    style["font"]
                )

                cell.fill = copy(
                    style["fill"]
                )

                cell.border = copy(
                    style["border"]
                )

                cell.alignment = copy(
                    style["alignment"]
                )

                cell.protection = copy(
                    style[
                        "protection"
                    ]
                )

                cell.number_format = (
                    style[
                        "number_format"
                    ]
                )

        if template_row_height:
            worksheet.row_dimensions[
                next_row
            ].height = (
                template_row_height
            )

        next_row += 1

    # ---------------------------------
    # 11. Append 36 Japan programs
    # ---------------------------------

    for program in japan_rows:
        for (
            column_number,
            header,
        ) in enumerate(
            EXPECTED_HEADERS,
            start=1,
        ):

            cell = worksheet.cell(
                row=next_row,
                column=column_number,
            )

            cell.value = convert_value(
                header,
                program.get(
                    header,
                    "",
                ),
            )

            if (
                column_number
                in template_styles
            ):
                style = template_styles[
                    column_number
                ]

                cell.font = copy(
                    style["font"]
                )

                cell.fill = copy(
                    style["fill"]
                )

                cell.border = copy(
                    style["border"]
                )

                cell.alignment = copy(
                    style["alignment"]
                )

                cell.protection = copy(
                    style[
                        "protection"
                    ]
                )

                cell.number_format = (
                    style[
                        "number_format"
                    ]
                )

            if header in {
                "application_deadline",
                "collected_at",
                "last_verified_at",
            }:
                cell.number_format = (
                    "yyyy-mm-dd"
                )

        if template_row_height:
            worksheet.row_dimensions[
                next_row
            ].height = (
                template_row_height
            )

        next_row += 1

    # ---------------------------------
    # 12. Update table / filter range
    # ---------------------------------

    final_row = next_row - 1

    final_range = (
        f"A1:U{final_row}"
    )

    for table in (
        worksheet.tables.values()
    ):
        table.ref = final_range

    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = (
            final_range
        )

    # ---------------------------------
    # 13. Save workbook
    # ---------------------------------

    workbook.save(
        WORKBOOK_PATH
    )

    workbook.close()

    # ---------------------------------
    # 14. Reload and verify
    # ---------------------------------

    verification_workbook = (
        openpyxl.load_workbook(
            WORKBOOK_PATH,
            read_only=True,
            data_only=True,
        )
    )

    verification_sheet = (
        verification_workbook[
            SHEET_NAME
        ]
    )

    japan_count = 0
    total_count = 0
    saved_program_ids = []

    for values in (
        verification_sheet.iter_rows(
            min_row=2,
            max_col=len(
                EXPECTED_HEADERS
            ),
            values_only=True,
        )
    ):
        if not any(
            value is not None
            for value in values
        ):
            continue

        total_count += 1

        program_id = values[0]
        university_id = values[1]

        if program_id is not None:
            saved_program_ids.append(
                str(program_id)
            )

        if is_japan_program(
            program_id,
            university_id,
        ):
            japan_count += 1

    verification_workbook.close()

    # ---------------------------------
    # 15. Verification rules
    # ---------------------------------

    if (
        japan_count
        != EXPECTED_JAPAN_PROGRAMS
    ):
        raise ValueError(
            "Verification failed: "
            f"expected 36 Japan programs, "
            f"found {japan_count}."
        )

    if (
        len(saved_program_ids)
        != len(set(saved_program_ids))
    ):
        raise ValueError(
            "Verification failed: "
            "duplicate program IDs exist "
            "in the workbook."
        )

    missing_ids = (
        set(source_program_ids)
        - set(saved_program_ids)
    )

    if missing_ids:
        raise ValueError(
            "Verification failed: "
            "some Japan program IDs "
            "were not saved: "
            + ", ".join(
                sorted(missing_ids)
            )
        )

    # ---------------------------------
    # 16. Summary
    # ---------------------------------

    print()
    print(
        "=== Japan Program Bulk "
        "Load Complete ==="
    )

    print(
        f"Japan programs inserted: "
        f"{japan_count}"
    )

    print(
        f"Non-Japan programs preserved: "
        f"{len(existing_non_japan_rows)}"
    )

    print(
        f"Total programs in workbook: "
        f"{total_count}"
    )

    print(
        f"Workbook: "
        f"{WORKBOOK_PATH}"
    )

    print(
        f"Safety backup: "
        f"{backup_path}"
    )

    print(
        "Verification: PASSED"
    )


if __name__ == "__main__":
    main()