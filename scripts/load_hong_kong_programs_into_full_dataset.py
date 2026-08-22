import argparse
import csv
import shutil
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl


CSV_PATH = Path(
    "data/cleaned/"
    "hong_kong_programs_final_ready.csv"
)

WORKBOOK_PATH = Path(
    "data/sample/"
    "06_full_dataset.xlsx"
)

BACKUP_DIR = Path(
    "data/backups/"
    "step_169_2bz"
)

SHEET_NAME = "programs"

EXPECTED_COUNT = 45


EXPECTED_HEADERS = [
    "program_id",
    "university_id",
    "program_name",
    "field_of_study",
    "degree_level",
    "duration_years",
    "study_mode",
    "language_of_instruction",
    "tuition_fee",
    "tuition_currency",
    "tuition_period",
    "minimum_gpa",
    "gpa_scale",
    "ielts_requirement",
    "toefl_requirement",
    "intake",
    "application_deadline",
    "program_url",
    "collected_at",
    "last_verified_at",
    "freshness_status",
]


EXPECTED_IDS = {
    f"prog_hk_{i:03d}"
    for i in range(1, 46)
}


NUMERIC_FIELDS = {
    "duration_years",
    "tuition_fee",
    "minimum_gpa",
    "gpa_scale",
    "ielts_requirement",
    "toefl_requirement",
}


DATE_FIELDS = {
    "application_deadline",
    "collected_at",
    "last_verified_at",
}


def clean(value):
    return str(value or "").strip()


def convert_value(field, value):

    value = clean(value)

    if value == "":
        return None

    if field in NUMERIC_FIELDS:

        number = float(value)

        if number.is_integer():
            return int(number)

        return number

    if field in DATE_FIELDS:
        return date.fromisoformat(value)

    return value


def canonical(value):

    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, float):

        if value.is_integer():
            return str(int(value))

        return format(value, ".15g")

    if isinstance(value, int):
        return str(value)

    return str(value).strip()


def is_hong_kong(program_id, university_id):

    return (
        clean(program_id).startswith("prog_hk_")
        or
        clean(university_id).startswith("uni_hk_")
    )


def read_source():

    if not CSV_PATH.exists():
        raise FileNotFoundError(
            f"Final Hong Kong CSV not found: {CSV_PATH}"
        )

    with CSV_PATH.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as file:

        reader = csv.DictReader(file)

        headers = reader.fieldnames or []
        rows = list(reader)

    if headers != EXPECTED_HEADERS:
        raise ValueError(
            "Hong Kong final CSV does not match "
            "the exact 21-column programme contract."
        )

    if len(rows) != EXPECTED_COUNT:
        raise ValueError(
            f"Expected {EXPECTED_COUNT} Hong Kong rows, "
            f"found {len(rows)}."
        )

    ids = [
        clean(row["program_id"])
        for row in rows
    ]

    if len(ids) != len(set(ids)):
        raise ValueError(
            "Duplicate program_id exists in Hong Kong CSV."
        )

    if set(ids) != EXPECTED_IDS:
        raise ValueError(
            "Hong Kong CSV ID set must be exactly "
            "prog_hk_001 through prog_hk_045."
        )

    for row in rows:

        program_id = clean(
            row["program_id"]
        )

        university_id = clean(
            row["university_id"]
        )

        if not program_id.startswith(
            "prog_hk_"
        ):
            raise ValueError(
                f"{program_id}: invalid Hong Kong program ID."
            )

        if not university_id.startswith(
            "uni_hk_"
        ):
            raise ValueError(
                f"{program_id}: invalid Hong Kong university ID."
            )

        if clean(
            row["freshness_status"]
        ) != "current":
            raise ValueError(
                f"{program_id}: freshness_status "
                "must be current."
            )

    return rows


def inspect_workbook():

    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {WORKBOOK_PATH}"
        )

    workbook = openpyxl.load_workbook(
        WORKBOOK_PATH,
        read_only=True,
        data_only=False,
    )

    try:

        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Sheet {SHEET_NAME!r} not found."
            )

        sheet = workbook[
            SHEET_NAME
        ]

        headers = [
            sheet.cell(
                row=1,
                column=i,
            ).value
            for i in range(
                1,
                len(EXPECTED_HEADERS) + 1,
            )
        ]

        headers = [
            clean(value)
            for value in headers
        ]

        if headers != EXPECTED_HEADERS:
            raise ValueError(
                "Workbook programs sheet does not "
                "match the exact 21-column schema."
            )

        all_program_ids = []
        hk_ids = []
        non_hk_ids = []

        nonempty_rows = 0

        for values in sheet.iter_rows(
            min_row=2,
            max_col=len(EXPECTED_HEADERS),
            values_only=True,
        ):

            if not any(
                value is not None
                and clean(value) != ""
                for value in values
            ):
                continue

            nonempty_rows += 1

            program_id = clean(
                values[0]
            )

            university_id = clean(
                values[1]
            )

            if not program_id:
                raise ValueError(
                    "Workbook contains a non-empty "
                    "programme row without program_id."
                )

            all_program_ids.append(
                program_id
            )

            if is_hong_kong(
                program_id,
                university_id,
            ):
                hk_ids.append(
                    program_id
                )
            else:
                non_hk_ids.append(
                    program_id
                )

        if (
            len(all_program_ids)
            != len(set(all_program_ids))
        ):
            raise ValueError(
                "Workbook currently contains duplicate "
                "program_id values."
            )

        return {
            "total": nonempty_rows,
            "hong_kong_ids": hk_ids,
            "non_hong_kong_ids": non_hk_ids,
        }

    finally:
        workbook.close()


def copy_row_style(
    sheet,
    source_row,
    target_row,
):

    if source_row is None:
        return

    for column in range(
        1,
        len(EXPECTED_HEADERS) + 1,
    ):

        source = sheet.cell(
            row=source_row,
            column=column,
        )

        target = sheet.cell(
            row=target_row,
            column=column,
        )

        if source.has_style:

            target.font = copy(
                source.font
            )

            target.fill = copy(
                source.fill
            )

            target.border = copy(
                source.border
            )

            target.alignment = copy(
                source.alignment
            )

            target.number_format = (
                source.number_format
            )

            target.protection = copy(
                source.protection
            )


def apply_to_workbook(
    source_rows,
    before_info,
):

    BACKUP_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    backup_path = BACKUP_DIR / (
        "06_full_dataset_before_"
        f"hong_kong_program_load_{timestamp}.xlsx"
    )

    shutil.copy2(
        WORKBOOK_PATH,
        backup_path,
    )

    workbook = openpyxl.load_workbook(
        WORKBOOK_PATH
    )

    try:

        sheet = workbook[
            SHEET_NAME
        ]

        hk_row_numbers = []

        template_row = None

        for row_number in range(
            2,
            sheet.max_row + 1,
        ):

            program_id = clean(
                sheet.cell(
                    row=row_number,
                    column=1,
                ).value
            )

            university_id = clean(
                sheet.cell(
                    row=row_number,
                    column=2,
                ).value
            )

            if (
                not program_id
                and not university_id
            ):
                continue

            if is_hong_kong(
                program_id,
                university_id,
            ):

                hk_row_numbers.append(
                    row_number
                )

            elif template_row is None:
                template_row = row_number

        for row_number in sorted(
            hk_row_numbers,
            reverse=True,
        ):

            sheet.delete_rows(
                row_number,
                1,
            )

        for source_row in source_rows:

            values = [
                convert_value(
                    field,
                    source_row[field],
                )
                for field in EXPECTED_HEADERS
            ]

            sheet.append(
                values
            )

            copy_row_style(
                sheet,
                template_row,
                sheet.max_row,
            )

        workbook.save(
            WORKBOOK_PATH
        )

    finally:
        workbook.close()

    return backup_path


def verify_after_save(
    source_rows,
    before_info,
):

    workbook = openpyxl.load_workbook(
        WORKBOOK_PATH,
        read_only=True,
        data_only=False,
    )

    try:

        sheet = workbook[
            SHEET_NAME
        ]

        saved_by_id = {}
        non_hk_ids = set()
        hk_ids = set()

        total = 0

        for values in sheet.iter_rows(
            min_row=2,
            max_col=len(EXPECTED_HEADERS),
            values_only=True,
        ):

            if not any(
                value is not None
                and clean(value) != ""
                for value in values
            ):
                continue

            total += 1

            program_id = clean(
                values[0]
            )

            university_id = clean(
                values[1]
            )

            if program_id in saved_by_id:
                raise ValueError(
                    f"Duplicate program_id after save: "
                    f"{program_id}"
                )

            saved_by_id[
                program_id
            ] = values

            if is_hong_kong(
                program_id,
                university_id,
            ):
                hk_ids.add(
                    program_id
                )
            else:
                non_hk_ids.add(
                    program_id
                )

        expected_non_hk = set(
            before_info[
                "non_hong_kong_ids"
            ]
        )

        if non_hk_ids != expected_non_hk:
            raise ValueError(
                "Non-Hong-Kong programme set changed."
            )

        if hk_ids != EXPECTED_IDS:
            raise ValueError(
                "Saved Hong Kong programme ID set mismatch."
            )

        if total != (
            len(expected_non_hk)
            + EXPECTED_COUNT
        ):
            raise ValueError(
                "Unexpected total programme count "
                "after workbook save."
            )

        mismatches = []

        for source_row in source_rows:

            program_id = clean(
                source_row["program_id"]
            )

            saved_values = saved_by_id[
                program_id
            ]

            for index, field in enumerate(
                EXPECTED_HEADERS
            ):

                expected_value = canonical(
                    convert_value(
                        field,
                        source_row[field],
                    )
                )

                actual_value = canonical(
                    saved_values[index]
                )

                if (
                    expected_value
                    != actual_value
                ):

                    mismatches.append(
                        (
                            program_id,
                            field,
                            expected_value,
                            actual_value,
                        )
                    )

        if mismatches:

            print()
            print(
                "First workbook value mismatches:"
            )

            for mismatch in mismatches[:10]:
                print(
                    mismatch
                )

            raise ValueError(
                "Hong Kong source/workbook "
                "value parity failed."
            )

        return {
            "total": total,
            "hong_kong": len(hk_ids),
            "non_hong_kong": len(
                non_hk_ids
            ),
            "mismatches": 0,
        }

    finally:
        workbook.close()


def main():

    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--apply",
        action="store_true",
        help=(
            "Actually modify the workbook. "
            "Without this flag the script performs "
            "a dry run only."
        ),
    )

    args = parser.parse_args()


    print("=" * 100)
    print(
        "STEP 169.2BZ - HONG KONG "
        "WORKBOOK INTEGRATION"
    )
    print("=" * 100)


    source_rows = read_source()

    before = inspect_workbook()


    print(
        "Source final-ready rows          :",
        len(source_rows),
    )

    print(
        "Workbook programmes before       :",
        before["total"],
    )

    print(
        "Existing Hong Kong rows          :",
        len(
            before["hong_kong_ids"]
        ),
    )

    print(
        "Non-Hong-Kong rows to preserve   :",
        len(
            before[
                "non_hong_kong_ids"
            ]
        ),
    )

    print(
        "Expected programmes after        :",
        len(
            before[
                "non_hong_kong_ids"
            ]
        )
        + EXPECTED_COUNT,
    )


    if not args.apply:

        print()
        print("=" * 100)

        print(
            "STEP 169.2BZ DRY RUN: PASS"
        )

        print(
            "SOURCE + WORKBOOK ARE READY "
            "FOR HONG KONG INTEGRATION"
        )

        print(
            "NO WORKBOOK OR MONGODB DATA "
            "WAS MODIFIED"
        )

        print()
        print(
            "Next command:"
        )

        print(
            "python "
            ".\\scripts\\"
            "load_hong_kong_programs_into_full_dataset.py "
            "--apply"
        )

        print("=" * 100)

        return


    backup_path = apply_to_workbook(
        source_rows,
        before,
    )


    after = verify_after_save(
        source_rows,
        before,
    )


    print()
    print(
        "Safety backup                    :",
        backup_path,
    )

    print(
        "Hong Kong programmes saved       :",
        after["hong_kong"],
    )

    print(
        "Non-Hong-Kong programmes kept    :",
        after["non_hong_kong"],
    )

    print(
        "Total programmes after           :",
        after["total"],
    )

    print(
        "HK source/workbook mismatches    :",
        after["mismatches"],
    )


    print()
    print("=" * 100)

    print(
        "STEP 169.2BZ WORKBOOK "
        "INTEGRATION: PASS"
    )

    print(
        "45 FINAL-READY HONG KONG "
        "PROGRAMMES ARE IN THE WORKBOOK"
    )

    print(
        "NON-HONG-KONG PROGRAMMES "
        "WERE PRESERVED"
    )

    print(
        "MONGODB WAS NOT MODIFIED"
    )

    print("=" * 100)


if __name__ == "__main__":
    main()
