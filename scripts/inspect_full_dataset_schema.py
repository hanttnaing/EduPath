from pathlib import Path

import openpyxl


WORKBOOK_PATH = Path(
    "data/sample/06_full_dataset.xlsx"
)


def main() -> None:
    workbook = openpyxl.load_workbook(
        WORKBOOK_PATH,
        read_only=True,
        data_only=True,
    )

    print("\n=== EduPath Full Dataset Schema ===\n")

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        headers = [
            cell.value
            for cell in next(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=1,
                )
            )
        ]

        print(f"Sheet: {sheet_name}")
        print(f"Columns ({len(headers)}):")

        for index, header in enumerate(
            headers,
            start=1,
        ):
            print(f"  {index}. {header}")

        print("-" * 60)


if __name__ == "__main__":
    main()