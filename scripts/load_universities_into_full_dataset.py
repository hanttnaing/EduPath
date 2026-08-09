import csv
import shutil
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl


CSV_PATH = Path(
    "data/cleaned/universities_master_ready.csv"
)

WORKBOOK_PATH = Path(
    "data/sample/06_full_dataset.xlsx"
)

SHEET_NAME = "universities"

EXPECTED_HEADERS = [
    "university_id",
    "university_name",
    "country_id",
    "city",
    "university_type",
    "official_website",
    "establishment_year",
    "global_ranking",
    "ranking_source",
    "ranking_year",
    "degree_levels",
    "scholarship_available",
    "source_url",
    "collected_at",
    "last_verified_at",
    "freshness_status",
]


def convert_value(
    header: str,
    value: str,
):
    value = value.strip()

    if value == "":
        return None

    if header in {
        "establishment_year",
        "global_ranking",
        "ranking_year",
    }:
        if value.isdigit():
            return int(value)

    if header in {
        "collected_at",
        "last_verified_at",
    }:
        try:
            return date.fromisoformat(value)
        except ValueError:
            return value

    return value


def main() -> None:
    # ---------------------------------
    # 1. Check required files
    # ---------------------------------

    if not CSV_PATH.exists():
        raise FileNotFoundError(
            f"CSV file not found: {CSV_PATH}"
        )

    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {WORKBOOK_PATH}"
        )

    # ---------------------------------
    # 2. Read university CSV
    # ---------------------------------

    with CSV_PATH.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as csv_file:
        reader = csv.DictReader(csv_file)

        csv_headers = reader.fieldnames

        if csv_headers != EXPECTED_HEADERS:
            raise ValueError(
                "CSV headers do not match "
                "the expected university schema."
            )

        university_rows = list(reader)

    print(
        f"University CSV records: "
        f"{len(university_rows)}"
    )

    if len(university_rows) != 50:
        raise ValueError(
            "Expected exactly 50 "
            "university records."
        )

    # ---------------------------------
    # 3. Create safety backup
    # ---------------------------------

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    backup_path = WORKBOOK_PATH.with_name(
        "06_full_dataset_before_"
        f"university_load_{timestamp}.xlsx"
    )

    shutil.copy2(
        WORKBOOK_PATH,
        backup_path,
    )

    print(
        f"Backup created: {backup_path}"
    )

    # ---------------------------------
    # 4. Open workbook
    # ---------------------------------

    workbook = openpyxl.load_workbook(
        WORKBOOK_PATH
    )

    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()

        raise ValueError(
            f"Sheet '{SHEET_NAME}' "
            "was not found."
        )

    worksheet = workbook[SHEET_NAME]

    # ---------------------------------
    # 5. Verify Excel headers
    # ---------------------------------

    workbook_headers = [
        worksheet.cell(
            row=1,
            column=column_number,
        ).value
        for column_number in range(
            1,
            len(EXPECTED_HEADERS) + 1,
        )
    ]

    if workbook_headers != EXPECTED_HEADERS:
        workbook.close()

        raise ValueError(
            "Universities sheet headers "
            "do not match the CSV schema."
        )

    print(
        "Workbook university schema: OK"
    )

    # ---------------------------------
    # 6. Preserve existing row style
    # ---------------------------------

    template_styles = {}

    if worksheet.max_row >= 2:
        for column_number in range(
            1,
            len(EXPECTED_HEADERS) + 1,
        ):
            cell = worksheet.cell(
                row=2,
                column=column_number,
            )

            template_styles[column_number] = {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(
                    cell.alignment
                ),
                "protection": copy(
                    cell.protection
                ),
                "number_format": (
                    cell.number_format
                ),
            }

    template_row_height = (
        worksheet.row_dimensions[2].height
        if worksheet.max_row >= 2
        else None
    )

    # ---------------------------------
    # 7. Remove old sample university
    #    rows only
    # ---------------------------------

    if worksheet.max_row > 1:
        worksheet.delete_rows(
            2,
            worksheet.max_row - 1,
        )

    # ---------------------------------
    # 8. Insert the 50 universities
    # ---------------------------------

    for row_number, university in enumerate(
        university_rows,
        start=2,
    ):
        for column_number, header in enumerate(
            EXPECTED_HEADERS,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
            )

            cell.value = convert_value(
                header,
                university.get(
                    header,
                    "",
                ),
            )

            if column_number in template_styles:
                style = template_styles[
                    column_number
                ]

                cell.font = copy(
                    style["font"]
                )
                cell.fill = copy(
                    style["fill"]
                )
                cell.border = copy(
                    style["border"]
                )
                cell.alignment = copy(
                    style["alignment"]
                )
                cell.protection = copy(
                    style["protection"]
                )
                cell.number_format = (
                    style["number_format"]
                )

            if header in {
                "collected_at",
                "last_verified_at",
            }:
                cell.number_format = (
                    "yyyy-mm-dd"
                )

        if template_row_height:
            worksheet.row_dimensions[
                row_number
            ].height = template_row_height

    # ---------------------------------
    # 9. Update table/filter range
    #    if the sheet uses them
    # ---------------------------------

    final_row = len(
        university_rows
    ) + 1

    final_range = (
        f"A1:P{final_row}"
    )

    for table in worksheet.tables.values():
        table.ref = final_range

    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = (
            final_range
        )

    # ---------------------------------
    # 10. Save workbook
    # ---------------------------------

    workbook.save(
        WORKBOOK_PATH
    )

    workbook.close()

    # ---------------------------------
    # 11. Verify saved workbook
    # ---------------------------------

    verification_workbook = (
        openpyxl.load_workbook(
            WORKBOOK_PATH,
            read_only=True,
            data_only=True,
        )
    )

    verification_sheet = (
        verification_workbook[
            SHEET_NAME
        ]
    )

    inserted_count = (
        verification_sheet.max_row - 1
    )

    verification_workbook.close()

    print()
    print(
        "=== University Bulk Load Complete ==="
    )
    print(
        f"Inserted universities: "
        f"{inserted_count}"
    )
    print(
        f"Workbook: {WORKBOOK_PATH}"
    )
    print(
        f"Safety backup: {backup_path}"
    )

    if inserted_count != 50:
        raise ValueError(
            "Verification failed. "
            "Workbook does not contain "
            "50 university rows."
        )

    print(
        "Verification: PASSED"
    )


if __name__ == "__main__":
    main()